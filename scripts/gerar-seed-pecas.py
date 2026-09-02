"""
Le o XLSX exportado do sistema legado GMAIS (Relatorios_GMAIS_XXXX.XLSX) e gera
db/seed-pecas.json com todas as pecas normalizadas prontas pra inserir na tabela pecas.

Uso: python scripts/gerar-seed-pecas.py <caminho_para_xlsx>

O JSON sai gravado em db/seed-pecas.json e vai pro repositorio. Em producao,
o docker-entrypoint.sh chama db/seed-pecas.ts que le esse JSON e insere
idempotentemente (ON CONFLICT (codigo) DO NOTHING).

Estrutura esperada do XLSX (aba 'Plan1'):
  Linhas 1-3: cabecalho / metadados
  Linha 4:    headers das colunas
  Linhas 5+:  dados
Colunas usadas (1-indexed):
  1  Codigo
  2  Descricao
  3  Unidade
  5  Nome Familia
  12 Cod Fab Original
  13 Cod Paralelo
  21 Saldo Fisico Total
  22 Ativo?
"""
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OUTPUT = ROOT / "db" / "seed-pecas.json"

# Unidades que aceitam fracao no dia-a-dia do almoxarifado (usado so na UI, mas
# util aqui pra normalizar valores duvidosos)
UNIDADES_VALIDAS = {
    "UN", "PC", "PÇ", "MT", "KG", "JG", "LT", "L", "KT", "PR", "BR",
    "CJ", "SC", "ML", "BT", "PA", "PE", "PT", "RL", "SV", "GL", "BL",
    "M3",
}


def limpa(v, maxlen=None):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    if maxlen and len(s) > maxlen:
        s = s[:maxlen]
    return s


def normaliza_unidade(v):
    if v is None:
        return "un"
    s = str(v).strip().upper()
    if not s or s == "1":
        return "UN"
    return s[:16]


def normaliza_saldo(v):
    if v is None:
        return "0"
    try:
        n = float(v)
        if n < 0:
            n = 0
        return f"{round(n, 3):.3f}"
    except (TypeError, ValueError):
        return "0"


def main():
    if len(sys.argv) != 2:
        print("Uso: python scripts/gerar-seed-pecas.py <caminho_para_xlsx>")
        sys.exit(2)

    xlsx = Path(sys.argv[1])
    print(f"[gerar-seed-pecas] Lendo {xlsx.name}…")
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["Plan1"]

    registros = []
    pulados_sem_cod = 0
    pulados_sem_desc = 0
    unidades_normalizadas = 0

    for row in range(5, ws.max_row + 1):
        codigo = limpa(ws.cell(row=row, column=1).value, maxlen=64)
        descricao = limpa(ws.cell(row=row, column=2).value, maxlen=255)
        if not codigo:
            pulados_sem_cod += 1
            continue
        if not descricao:
            pulados_sem_desc += 1
            continue

        unidade_bruta = ws.cell(row=row, column=3).value
        unidade = normaliza_unidade(unidade_bruta)
        if str(unidade_bruta or "").strip() != unidade and str(unidade_bruta or "").strip() != "":
            unidades_normalizadas += 1

        registros.append(
            {
                "codigo": codigo,
                "nome": descricao,
                "unidade": unidade,
                "saldo": normaliza_saldo(ws.cell(row=row, column=21).value),
                "familia": limpa(ws.cell(row=row, column=5).value, maxlen=64),
                "codigoFabricante": limpa(ws.cell(row=row, column=12).value, maxlen=64),
                "codigoParalelo": limpa(ws.cell(row=row, column=13).value, maxlen=64),
            }
        )

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT.open("w", encoding="utf-8") as f:
        json.dump(registros, f, ensure_ascii=False)

    com_saldo = sum(1 for r in registros if float(r["saldo"]) > 0)
    com_familia = sum(1 for r in registros if r["familia"])
    com_cod_fab = sum(1 for r in registros if r["codigoFabricante"])

    print(f"[gerar-seed-pecas] {len(registros)} peças gravadas em {OUTPUT}")
    print(f"[gerar-seed-pecas]   com saldo > 0:        {com_saldo}")
    print(f"[gerar-seed-pecas]   com família:          {com_familia}")
    print(f"[gerar-seed-pecas]   com cod. fabricante:  {com_cod_fab}")
    print(f"[gerar-seed-pecas]   unidades normalizadas: {unidades_normalizadas}")
    print(f"[gerar-seed-pecas]   linhas puladas (sem cód): {pulados_sem_cod}")
    print(f"[gerar-seed-pecas]   linhas puladas (sem desc): {pulados_sem_desc}")


if __name__ == "__main__":
    main()
