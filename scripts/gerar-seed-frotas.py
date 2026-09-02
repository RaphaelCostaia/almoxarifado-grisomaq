"""
Le os dois XLSX enviados pelo cliente e gera db/seed-frotas.json com os
266 registros normalizados prontos pra inserir na tabela frotas.

Rodar uma vez: python scripts/gerar-seed-frotas.py <caminho_equipamentos.xlsx> <caminho_implementos.xlsx>

O JSON sai gravado em db/seed-frotas.json e vai pro repositorio.
Em producao, o docker-entrypoint.sh chama db/seed-frotas.ts que le esse JSON e insere.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OUTPUT = ROOT / "db" / "seed-frotas.json"


def limpa(v):
    """Normaliza célula: remove tracinhos placeholder, espacos duplos, strings vazias -> None."""
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    # Tracinhos placeholder do sistema legado
    if re.fullmatch(r"[-]{2,}", s):
        return None
    # "0" no ano e valores tipo isso viram None se numérico invalido
    return s


def limpa_ano(v):
    s = limpa(v)
    if s is None:
        return None
    try:
        n = int(float(s))
        if 1950 <= n <= 2100:
            return str(n)
    except ValueError:
        pass
    return None


def extrai_marca_do_modelo(modelo: str | None):
    """De '1-TOYOTA HILUX CD 4X4' extrai 'TOYOTA'. Retorna (marca, modelo_curto)."""
    if not modelo:
        return None, None
    # Remove prefixo tipo "1-" ou "587-"
    m = re.match(r"^\s*\d+-\s*(.+)$", modelo)
    limpo = (m.group(1) if m else modelo).strip()
    partes = limpo.split(" ", 1)
    marca = partes[0] if partes else None
    return marca, limpo


def carregar_equipamentos(caminho: Path):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb["Plan1"]
    registros = []
    # Cabecalho na linha 4, dados a partir de 5
    for row in range(5, ws.max_row + 1):
        numero = limpa(ws.cell(row=row, column=1).value)
        if not numero:
            continue
        modelo_bruto = limpa(ws.cell(row=row, column=2).value)
        marca, modelo_limpo = extrai_marca_do_modelo(modelo_bruto)
        em_operacao = str(ws.cell(row=row, column=20).value or "").strip().lower()
        data_baixa = limpa(ws.cell(row=row, column=21).value)
        ativo = 1 if em_operacao.startswith("s") and not data_baixa else 0
        registros.append(
            {
                "numero": str(numero).strip(),
                "categoria": "equipamento",
                "modelo": modelo_limpo,
                "marca": marca,
                "descricao": None,
                "ano": limpa_ano(ws.cell(row=row, column=7).value),
                "placa": limpa(ws.cell(row=row, column=8).value),
                "chassi": limpa(ws.cell(row=row, column=4).value),
                "localizacao": limpa(ws.cell(row=row, column=11).value),
                "proprietario": limpa(ws.cell(row=row, column=13).value),
                "ativo": ativo,
                "observacoes": limpa(ws.cell(row=row, column=23).value),
            }
        )
    return registros


def carregar_implementos(caminho: Path):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb["Plan1"]
    registros = []
    for row in range(5, ws.max_row + 1):
        numero = limpa(ws.cell(row=row, column=2).value)
        if not numero:
            continue
        modelo_bruto = limpa(ws.cell(row=row, column=3).value)
        _, modelo_limpo = extrai_marca_do_modelo(modelo_bruto)
        marca = limpa(ws.cell(row=row, column=11).value)
        data_baixa = limpa(ws.cell(row=row, column=33).value)
        ativo = 0 if data_baixa else 1
        registros.append(
            {
                "numero": str(numero).strip(),
                "categoria": "implemento",
                "modelo": modelo_limpo,
                "marca": marca,
                "descricao": limpa(ws.cell(row=row, column=4).value),
                "ano": limpa_ano(ws.cell(row=row, column=8).value),
                "placa": limpa(ws.cell(row=row, column=12).value),
                "chassi": limpa(ws.cell(row=row, column=13).value),
                "localizacao": limpa(ws.cell(row=row, column=18).value),
                "proprietario": limpa(ws.cell(row=row, column=19).value),
                "ativo": ativo,
                "observacoes": limpa(ws.cell(row=row, column=35).value),
            }
        )
    return registros


def main():
    if len(sys.argv) != 3:
        print("Uso: python scripts/gerar-seed-frotas.py <equipamentos.xlsx> <implementos.xlsx>")
        sys.exit(2)

    equip_path = Path(sys.argv[1])
    impl_path = Path(sys.argv[2])
    print(f"[gerar-seed] Lendo equipamentos: {equip_path.name}")
    equip = carregar_equipamentos(equip_path)
    print(f"[gerar-seed] Lendo implementos: {impl_path.name}")
    impl = carregar_implementos(impl_path)

    # Detecta duplicidade de numero — se houver, coloca sufixo pra nao colidir
    todos = []
    vistos = set()
    for r in equip + impl:
        num = r["numero"]
        chave = num
        n = 1
        while chave in vistos:
            n += 1
            chave = f"{num}-{n}"
        r["numero"] = chave
        vistos.add(chave)
        todos.append(r)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT.open("w", encoding="utf-8") as f:
        json.dump(todos, f, ensure_ascii=False, indent=2)
    print(f"[gerar-seed] {len(todos)} registros gravados em {OUTPUT}")
    print(f"[gerar-seed]   equipamentos: {sum(1 for r in todos if r['categoria']=='equipamento')}")
    print(f"[gerar-seed]   implementos: {sum(1 for r in todos if r['categoria']=='implemento')}")
    print(f"[gerar-seed]   ativos: {sum(1 for r in todos if r['ativo']==1)}")


if __name__ == "__main__":
    main()
